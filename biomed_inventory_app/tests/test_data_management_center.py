import asyncio
import io
import os
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import sessionmaker

from app.database import Base, build_engine
from app.data_management.template_registry import all_datasets, get_dataset, validate_registry
from app.erp_models import Client
from app.models.foundation import DataValidationError, ImportBatch, ImportRow
from app.routers import data_management_api as api


class DataManagementCenterTest(unittest.TestCase):
    def setUp(self):
        fd, self.db_path = tempfile.mkstemp(suffix=".db")
        os.close(fd)
        self.engine = build_engine(f"sqlite:///{self.db_path}")
        Base.metadata.create_all(self.engine)
        self.Session = sessionmaker(bind=self.engine, future=True)
        self.tempdir = tempfile.TemporaryDirectory()
        self.old_session = api.SessionLocal
        self.old_upload_root = api.UPLOAD_ROOT
        self.old_export_root = api.EXPORT_ROOT
        self.old_template_root = api.TEMPLATE_ROOT
        api.SessionLocal = self.Session
        api.UPLOAD_ROOT = Path(self.tempdir.name) / "imports"
        api.EXPORT_ROOT = Path(self.tempdir.name) / "exports"
        api.TEMPLATE_ROOT = Path(self.tempdir.name) / "templates"
        for path in [api.UPLOAD_ROOT / "original", api.UPLOAD_ROOT / "processed", api.UPLOAD_ROOT / "error-reports", api.EXPORT_ROOT, api.TEMPLATE_ROOT]:
            path.mkdir(parents=True, exist_ok=True)

    def tearDown(self):
        api.SessionLocal = self.old_session
        api.UPLOAD_ROOT = self.old_upload_root
        api.EXPORT_ROOT = self.old_export_root
        api.TEMPLATE_ROOT = self.old_template_root
        self.engine.dispose()
        self.tempdir.cleanup()
        if os.path.exists(self.db_path):
            os.unlink(self.db_path)

    def _request(self):
        return SimpleNamespace(client=SimpleNamespace(host="127.0.0.1"), headers={"user-agent": "unit-test"})

    def _upload(self, name, content):
        return UploadFile(filename=name, file=io.BytesIO(content))

    def test_registry_loads_without_duplicate_fields(self):
        validate_registry()
        datasets = all_datasets()
        self.assertGreaterEqual(len(datasets), 14)
        self.assertIn("client_name", get_dataset("clients").required_fields)

    def test_template_generation_has_expected_sheets_and_empty_data_sheet(self):
        workbook = api._template_workbook(get_dataset("equipment"))
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        loaded = load_workbook(output)
        self.assertEqual(loaded.sheetnames, ["Data", "Instructions", "Accepted Values", "Example Data"])
        data = loaded["Data"]
        self.assertEqual(data.max_row, 1)
        self.assertIn("serial_number", [cell.value for cell in data[1]])

    def test_unsupported_and_oversized_uploads_are_rejected(self):
        with self.assertRaises(Exception):
            api._read_upload(self._upload("../bad.exe", b"not ok"), b"not ok")
        with self.assertRaises(Exception):
            api._read_upload(self._upload("clients.csv", b"a\n" + b"x" * (api.MAX_UPLOAD_BYTES + 1)), b"a\n" + b"x" * (api.MAX_UPLOAD_BYTES + 1))

    def test_import_stages_rows_and_does_not_write_production_table(self):
        content = b"Hospital,City\nExample Hospital,Beirut\n"
        result = asyncio.run(
            api.create_import(
                self._request(),
                dataset_key="clients",
                import_mode="validate_only",
                mapping_json=None,
                notes="unit",
                file=self._upload("../../clients.csv", content),
            )
        )
        self.assertEqual(result["summary"]["ready_rows"], 1)
        with self.Session() as db:
            self.assertEqual(db.query(ImportBatch).count(), 1)
            self.assertEqual(db.query(ImportRow).count(), 1)
            self.assertEqual(db.query(DataValidationError).count(), 0)
            self.assertEqual(db.query(Client).count(), 0)
        stored = list((api.UPLOAD_ROOT / "original").glob("*"))
        self.assertEqual(len(stored), 1)
        self.assertNotIn("..", stored[0].name)

    def test_blocking_validation_error_prevents_confirmation(self):
        content = b"City\nBeirut\n"
        result = asyncio.run(
            api.create_import(
                self._request(),
                dataset_key="clients",
                import_mode="validate_only",
                mapping_json=None,
                notes=None,
                file=self._upload("clients.csv", content),
            )
        )
        self.assertEqual(result["summary"]["error_rows"], 1)
        with self.assertRaises(Exception):
            api.confirm_import(result["batch_id"])

    def test_export_rejects_unsupported_columns_and_exports_empty_csv(self):
        with self.assertRaises(Exception):
            api.export_preview(api.ExportPreviewRequest(dataset_key="clients", columns=["password_hash"]))
        response = api.export_download(api.ExportDownloadRequest(dataset_key="clients", columns=["client_name"], format="csv"), self._request())
        self.assertEqual(response.media_type, "text/csv")

    def test_data_quality_counts_fixture_data(self):
        with self.Session() as db:
            db.add(ImportBatch(source_type="clients", source_filename="bad.csv", status="validation_failed", total_rows=1, failed_rows=1))
            db.flush()
            batch_id = db.query(ImportBatch).one().id
            db.add(DataValidationError(import_batch_id=batch_id, error_code="required", error_message="client_name is required", severity="error"))
            db.commit()
            metrics = {row["metric_key"]: row["count"] for row in api.data_quality(db)}
        self.assertEqual(metrics["unresolved_validation_errors"], 1)

    def test_routes_are_registered(self):
        import app.main as main

        paths = {route.path for route in main.app.routes if hasattr(route, "path")}
        self.assertIn("/administration/data-management", paths)
        self.assertIn("/api/data-management/summary", paths)
        self.assertIn("/api/data-management/imports", paths)


if __name__ == "__main__":
    unittest.main()
