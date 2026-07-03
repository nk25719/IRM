from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


COLUMN_ALIASES = {
    "item_code": {"code", "item code", "item_code", "ref", "reference", "pn", "part no", "part number", "part", "mpn"},
    "manufacturer_part_number": {"manufacturer part number", "manufacturer pn", "mfr pn", "mfg part", "model"},
    "description": {"description", "item", "item description", "product", "product description", "details"},
    "quantity": {"qty", "quantity", "qte", "count", "amount"},
    "unit_price": {"price", "unit price", "unit_price", "unit cost", "rate", "selling price"},
    "discount_percent": {"discount", "discount %", "discount percent", "disc", "disc %"},
    "warranty": {"warranty", "guarantee"},
    "delivery_time": {"delivery", "delivery time", "lead time", "eta"},
}


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[\s\-_./]+", " ", text)
    return text.strip()


def map_headers(headers: list[Any]) -> dict[int, str]:
    mapped: dict[int, str] = {}
    normalized = [normalize_header(header) for header in headers]
    for idx, header in enumerate(normalized):
        for field, aliases in COLUMN_ALIASES.items():
            header_words = set(header.split())
            if header in aliases or any(set(alias.split()).issubset(header_words) for alias in aliases if " " in alias):
                mapped[idx] = field
                break
    return mapped


def parse_number(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"[^0-9.\-]", "", str(value))
    try:
        return float(text) if text else default
    except ValueError:
        return default


def normalize_imported_item(raw: dict[str, Any], row_number: int) -> dict[str, Any]:
    quantity = parse_number(raw.get("quantity"), 1.0)
    unit_price = parse_number(raw.get("unit_price"), 0.0)
    discount_percent = parse_number(raw.get("discount_percent"), 0.0)
    line_total = round(quantity * unit_price * (1 - discount_percent / 100), 2)
    return {
        "row_number": row_number,
        "inventory_item_id": None,
        "item_code": str(raw.get("item_code") or "").strip() or None,
        "manufacturer_part_number": str(raw.get("manufacturer_part_number") or "").strip() or None,
        "description": str(raw.get("description") or raw.get("item_code") or "").strip(),
        "quantity": quantity,
        "unit_price": unit_price,
        "discount_percent": discount_percent,
        "line_total": line_total,
        "warranty": str(raw.get("warranty") or "").strip() or None,
        "delivery_time": str(raw.get("delivery_time") or "").strip() or None,
    }


def parse_excel_bytes(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    header_index = 0
    header_map: dict[int, str] = {}
    for idx, row in enumerate(rows[:8]):
        candidate = map_headers(list(row))
        if len(candidate) >= 2:
            header_index = idx
            header_map = candidate
            break
    if not header_map:
        header_map = map_headers(list(rows[0]))

    items = []
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        raw = {field: row[idx] for idx, field in header_map.items() if idx < len(row)}
        if not any(str(value or "").strip() for value in raw.values()):
            continue
        item = normalize_imported_item(raw, row_number)
        if item["description"] or item["item_code"]:
            items.append(item)
    return items


def parse_csv_text(text: str) -> list[dict[str, Any]]:
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []
    header_map = map_headers(rows[0])
    items = []
    for row_number, row in enumerate(rows[1:], start=2):
        raw = {field: row[idx] for idx, field in header_map.items() if idx < len(row)}
        if raw:
            items.append(normalize_imported_item(raw, row_number))
    return items


def extract_text_from_upload(filename: str, content: bytes) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix == ".pdf":
        try:
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(content))
            return "\n".join(page.extract_text() or "" for page in reader.pages)
        except Exception:
            return ""
    if suffix == ".docx":
        try:
            from docx import Document

            document = Document(io.BytesIO(content))
            return "\n".join(paragraph.text for paragraph in document.paragraphs)
        except Exception:
            return ""
    return content.decode("utf-8", "ignore")


def parse_text_items(text: str) -> list[dict[str, Any]]:
    items = []
    for row_number, line in enumerate(text.splitlines(), start=1):
        line = line.strip()
        if not line or len(line) < 3:
            continue
        parts = re.split(r"\s{2,}|\t|\|", line)
        if len(parts) < 2:
            continue
        maybe_qty = next((parse_number(part, -1) for part in parts if re.fullmatch(r"\d+(\.\d+)?", part.strip())), -1)
        maybe_price = next((parse_number(part, -1) for part in reversed(parts) if re.search(r"\d", part) and parse_number(part, -1) >= 0), 0)
        raw = {
            "item_code": parts[0],
            "description": parts[1] if len(parts) > 1 else line,
            "quantity": maybe_qty if maybe_qty > 0 else 1,
            "unit_price": maybe_price if maybe_price >= 0 else 0,
        }
        items.append(normalize_imported_item(raw, row_number))
    return items


def parse_upload(filename: str, content: bytes) -> dict[str, Any]:
    suffix = Path(filename).suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        items = parse_excel_bytes(content)
        source_type = "excel"
    elif suffix == ".csv":
        items = parse_csv_text(content.decode("utf-8", "ignore"))
        source_type = "csv"
    else:
        text = extract_text_from_upload(filename, content)
        items = parse_text_items(text)
        source_type = suffix.lstrip(".") or "text"
    return {"filename": filename, "source_type": source_type, "items": items, "item_count": len(items)}
