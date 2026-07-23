from __future__ import annotations

import csv
import io
import json
import os
import re
import shutil
import sqlite3
import subprocess
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import quote, urlparse, urlunparse

import pandas as pd
from fastapi import APIRouter, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, RedirectResponse, Response, StreamingResponse
from openpyxl import Workbook
from sqlalchemy import create_engine, inspect, text

from app import legacy_main
from app.config.database import get_database_url, get_sqlite_database_path, is_postgresql_database
from app.data_management.template_registry import all_datasets, get_dataset

DATABASE_URL = get_database_url()


router = APIRouter(tags=["Admin Foundation"])

BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
BACKUP_DIR = BASE_DIR / "backups"
BACKUP_DIR.mkdir(exist_ok=True)

DEFAULT_ROLES = {
    "admin": [
        "view_all_clients", "view_prices", "edit_quotations", "approve_quotations", "export_pdf",
        "import_data", "manage_users", "create_backup", "view_reports", "view_after_sales_cases",
        "edit_service_calls", "view_database_map", "run_select_queries", "data.templates.download",
        "data.import.preview", "data.import.execute", "data.export.basic", "data.export.financial",
        "data.import.history", "data.validation.resolve", "data_management.view",
        "data_management.templates.download", "data_management.import.preview",
        "data_management.import.execute", "data_management.export", "data_management.export_financial",
        "data_management.validation.view", "data_management.validation.resolve",
        "data_management.history.view", "service_intelligence.view", "service_intelligence.refresh",
        "service_intelligence.manage_opportunities", "service_intelligence.assign_opportunities",
        "service_intelligence.import", "service_intelligence.export", "service_intelligence.reconcile",
        "service_intelligence.configure", "manufacturer_agreements.view", "manufacturer_agreements.manage",
        "manufacturer_agreements.import", "manufacturer_agreements.reconcile", "manufacturer_agreements.view_values",
        "customer_contracts.view", "customer_contracts.manage", "customer_contracts.import",
        "customer_contracts.renew", "customer_contracts.view_values", "contract_pm.manage",
    ],
    "sales": ["view_all_clients", "view_prices", "edit_quotations", "export_pdf", "view_reports"],
    "procurement": ["view_all_clients", "view_prices", "view_reports"],
    "warehouse": ["view_all_clients", "view_reports"],
    "after_sales": ["view_all_clients", "view_after_sales_cases", "edit_service_calls", "view_reports", "service_intelligence.view", "service_intelligence.refresh", "service_intelligence.manage_opportunities", "service_intelligence.import", "service_intelligence.export", "service_intelligence.reconcile", "customer_contracts.view", "customer_contracts.manage", "customer_contracts.import", "customer_contracts.renew", "contract_pm.manage"],
    "engineer": ["view_after_sales_cases", "edit_service_calls"],
    "viewer": ["view_reports"],
}

PRICE_FIELDS = {
    "unit_price", "total_price", "line_total", "amount", "subtotal", "vat_amount", "total_amount",
    "supplier_price", "invoice_value", "credit_balance",
}

TABLE_EXPLANATIONS = {
    "clients": "Hospitals and customer organizations. Most ERM work starts by linking a case, equipment, quote, or order to a client.",
    "departments": "Customer departments inside a client, such as ICU, ER, OR, or biomedical engineering.",
    "contacts": "People at a client or department, with role, email, and phone details.",
    "equipment": "Installed equipment records, usually linked to a client, department, serial number, and model.",
    "service_calls": "After-Sales service visits and issues, including engineer assignment, dates, status, and resolution.",
    "cases": "Unified workflow records that connect requests, service, quotations, POs, deliveries, invoices, and follow-up.",
    "case_items": "Line-level needs inside a case, such as requested parts, equipment, quantities, and procurement state.",
    "inventory_items": "Warehouse item master records used for spare parts and stock lookup.",
    "quotations": "Financial offers or proposals sent to customers before confirmed orders.",
    "quotation_items": "Line items inside quotations, including quantity, item code, description, and prices.",
    "quotation_equipment_groups": "Equipment/serial-number groups inside service quotations; lets the same part appear under different serial numbers.",
    "purchase_orders": "Supplier purchase order headers.",
    "shipments": "Inbound supplier shipments connected to purchase orders and stock items.",
    "receptions": "Warehouse receiving records for shipments.",
    "delivery_orders": "Outbound delivery records to customers.",
    "invoices": "Customer billing documents and invoice status.",
}

IMPORT_TARGETS = {
    "clients": {
        "table": "clients",
        "fields": ["name", "city", "address", "main_contact", "contact_email", "phone", "status", "notes"],
        "required": ["name"],
        "match": "name",
    },
    "departments": {
        "table": "departments",
        "fields": ["client_name", "department_name", "floor_location", "main_contact_name", "phone", "email", "notes"],
        "required": ["client_name", "department_name"],
        "match": "client_department",
    },
    "contacts": {
        "table": "contacts",
        "fields": ["client_name", "department_name", "name", "role", "email", "phone", "notes"],
        "required": ["client_name", "name"],
        "match": "contact",
    },
    "equipment": {
        "table": "equipment",
        "fields": ["client_name", "department_name", "asset_tag", "serial_number", "manufacturer", "model", "status"],
        "required": ["serial_number"],
        "match": "serial_number",
    },
    "inventory_items": {
        "table": "inventory_items",
        "fields": ["pn", "item_code", "description", "device_family", "barcode", "active"],
        "required": ["pn"],
        "match": "pn",
    },
    "service_calls": {
        "table": "service_calls",
        "fields": ["client_name", "case_no", "source", "source_reference", "call_no", "status", "engineer", "issue", "opened_at", "notes"],
        "required": ["client_name", "call_no"],
        "match": "service_call",
    },
    "pm_tasks": {
        "table": "pm_tasks",
        "fields": ["asset_id", "task_name", "description", "status", "assigned_to", "due_date", "notes"],
        "required": ["task_name"],
        "match": "",
    },
    "quotation_items": {
        "table": "quotation_items",
        "fields": ["quotation_id", "item_code", "ref", "description", "quantity", "qty", "unit_price", "item_type", "notes"],
        "required": ["description"],
        "match": "",
    },
    "purchase_order_items": {
        "table": "purchase_order_items",
        "fields": ["po_no", "pn", "ref", "description", "qty", "location", "status", "notes"],
        "required": ["po_no", "description", "qty"],
        "match": "",
    },
}

DATA_MANAGEMENT_TYPES = {
    "clients": {
        "label": "Clients",
        "fields": ["client_code", "client_name", "city", "address", "main_contact", "contact_email", "phone", "status", "notes"],
        "required": ["client_name"],
        "accepted_values": {"status": ["active", "inactive", "archived"]},
        "example": {"client_code": "CLIENT-001", "client_name": "Example Hospital", "city": "Beirut", "status": "active"},
    },
    "departments": {
        "label": "Departments",
        "fields": ["client_code", "client_name", "department_code", "department_name", "floor_location", "main_contact_name", "phone", "email", "notes"],
        "required": ["client_name", "department_name"],
        "accepted_values": {},
        "example": {"client_code": "CLIENT-001", "department_code": "ICU", "department_name": "ICU"},
    },
    "contacts": {
        "label": "Contacts",
        "fields": ["client_code", "client_name", "department_name", "contact_name", "role", "email", "phone", "notes"],
        "required": ["client_name", "contact_name"],
        "accepted_values": {},
        "example": {"client_code": "CLIENT-001", "contact_name": "Jane Example", "role": "Biomedical Engineer"},
    },
    "equipment": {
        "label": "Equipment",
        "fields": [
            "client_code", "client_name", "site_code", "department_name", "equipment_category", "manufacturer",
            "model", "serial_number", "asset_number", "installation_date", "warranty_start_date",
            "warranty_end_date", "status", "location", "notes",
        ],
        "required": ["client_name", "serial_number"],
        "accepted_values": {"status": ["active", "inactive", "under_service", "retired"]},
        "example": {"client_code": "CLIENT-001", "equipment_category": "Patient Monitoring", "manufacturer": "GE Healthcare", "model": "Dash 4000", "serial_number": "SN-001", "status": "active"},
    },
    "equipment_models": {
        "label": "Equipment Models",
        "fields": ["manufacturer_code", "manufacturer", "equipment_category_code", "equipment_category", "model", "description", "status", "notes"],
        "required": ["model"],
        "accepted_values": {"status": ["active", "inactive"]},
        "example": {"manufacturer": "GE Healthcare", "equipment_category": "Patient Monitoring", "model": "Dash 4000", "status": "active"},
    },
    "manufacturers": {
        "label": "Manufacturers",
        "fields": ["manufacturer_code", "manufacturer_name", "legal_name", "website", "email", "phone", "country_code", "status", "notes"],
        "required": ["manufacturer_name"],
        "accepted_values": {"status": ["active", "inactive"]},
        "example": {"manufacturer_code": "GE", "manufacturer_name": "GE Healthcare", "status": "active"},
    },
    "suppliers": {
        "label": "Suppliers",
        "fields": ["supplier_code", "supplier_name", "legal_name", "email", "phone", "website", "tax_number", "country_code", "status", "notes"],
        "required": ["supplier_code", "supplier_name"],
        "accepted_values": {"status": ["active", "inactive"]},
        "example": {"supplier_code": "SUP-001", "supplier_name": "Example Supplier", "status": "active"},
    },
    "inventory_items": {
        "label": "Inventory Items",
        "fields": ["item_code", "pn", "description", "item_category", "device_family", "manufacturer", "model", "barcode", "status", "notes"],
        "required": ["pn", "description"],
        "accepted_values": {"status": ["active", "inactive"]},
        "example": {"pn": "PN-001", "description": "ECG Cable", "item_category": "spare_parts", "status": "active"},
    },
    "service_calls": {
        "label": "Service Calls",
        "fields": ["client_code", "client_name", "case_no", "call_no", "equipment_serial", "engineer", "issue", "opened_at", "status", "notes"],
        "required": ["client_name", "call_no"],
        "accepted_values": {"status": ["open", "in_progress", "closed", "cancelled"]},
        "example": {"client_name": "Example Hospital", "call_no": "SC-001", "issue": "No display", "status": "open"},
    },
    "preventive_maintenance": {
        "label": "Preventive Maintenance",
        "fields": ["client_code", "equipment_serial", "task_name", "due_date", "assigned_to", "status", "notes"],
        "required": ["task_name"],
        "accepted_values": {"status": ["planned", "due", "completed", "cancelled"]},
        "example": {"equipment_serial": "SN-001", "task_name": "Quarterly PM", "status": "planned"},
    },
    "contracts": {
        "label": "Contracts",
        "fields": ["client_code", "contract_number", "contract_type", "start_date", "end_date", "status", "notes"],
        "required": ["client_code", "contract_number"],
        "accepted_values": {"status": ["draft", "active", "expired", "cancelled"]},
        "example": {"client_code": "CLIENT-001", "contract_number": "CTR-001", "contract_type": "Service"},
    },
    "quotations": {
        "label": "Quotations",
        "fields": ["client_code", "quotation_number", "quotation_date", "valid_until", "status", "currency", "notes"],
        "required": ["client_code", "quotation_number"],
        "accepted_values": {"status": ["draft", "sent", "approved", "rejected"]},
        "example": {"client_code": "CLIENT-001", "quotation_number": "Q-001", "status": "draft", "currency": "USD"},
    },
}

REPORTS = {
    "open_service_calls_by_engineer": "SELECT engineer, status, call_no, issue, opened_at FROM service_calls WHERE lower(COALESCE(status,'')) NOT IN ('closed','completed','cancelled') ORDER BY engineer, opened_at DESC",
    "pending_quotations_by_salesperson": "SELECT sales_person, quotation_number, quotation_no, status, quotation_date, valid_until, total_amount, amount FROM quotations WHERE lower(COALESCE(status,'draft')) IN ('draft','pending','sent','reviewed') ORDER BY sales_person, quotation_date DESC",
    "pending_customer_orders": "SELECT co_no, client_order_no, status, order_date, expected_date, notes FROM customer_orders WHERE lower(COALESCE(status,'open')) NOT IN ('delivered','closed','cancelled') ORDER BY order_date DESC",
    "pending_pos": "SELECT po_no, supplier, status, po_date, expected_date, shipping_status, reception_status FROM purchase_orders WHERE lower(COALESCE(status,'open')) NOT IN ('received','closed','cancelled') ORDER BY po_date DESC",
    "upcoming_shipments": "SELECT shipment_no, supplier_id, status, shipment_date, expected_arrival, notes FROM shipments WHERE lower(COALESCE(status,'pending')) NOT IN ('received','closed','cancelled') ORDER BY expected_arrival",
    "warehouse_low_stock": "SELECT pn, description, location, system_qty, physical_qty, reserved_qty FROM inventory WHERE COALESCE(physical_qty, system_qty, 0) <= COALESCE(reserved_qty, 0) ORDER BY pn",
    "equipment_by_client": "SELECT c.name AS client, d.department_name, e.asset_tag, e.serial_number, e.manufacturer, e.model, e.status FROM equipment e LEFT JOIN clients c ON c.id=e.client_id LEFT JOIN departments d ON d.id=e.department_id ORDER BY c.name, d.department_name, e.model",
    "cases_by_status": "SELECT status, workflow_state, case_type, COUNT(*) AS case_count FROM cases GROUP BY status, workflow_state, case_type ORDER BY status, workflow_state",
    "pm_tasks_due_this_month": "SELECT task_name, assigned_to, status, due_date, notes FROM pm_tasks WHERE due_date >= date('now','start of month') AND due_date < date('now','start of month','+1 month') ORDER BY due_date",
}


def now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def db() -> sqlite3.Connection:
    return legacy_main.db()


def is_postgres_url(url: str = DATABASE_URL) -> bool:
    return is_postgresql_database(url)


def sqlite_path() -> Path:
    return get_sqlite_database_path(DATABASE_URL)


def masked_database_url() -> dict[str, Any]:
    if is_postgres_url():
        parsed = urlparse(DATABASE_URL)
        netloc = parsed.netloc
        if "@" in netloc:
            userinfo, host = netloc.rsplit("@", 1)
            username = userinfo.split(":", 1)[0]
            netloc = f"{username}:***@{host}"
        safe = urlunparse(parsed._replace(netloc=netloc, query=""))
        return {"engine": "PostgreSQL", "location_type": "DATABASE_URL", "safe_location": safe}
    path = sqlite_path()
    return {"engine": "SQLite", "location_type": "file", "safe_location": str(path)}


def role_from_request(request: Request) -> str:
    return str(request.session.get("role") or legacy_main.APP_ROLE or "viewer")


def username_from_request(request: Request) -> str:
    return str(request.session.get("username") or "system")


def ensure_admin_foundation() -> None:
    with db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS roles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                description TEXT,
                created_at TEXT,
                updated_at TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS permissions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                description TEXT,
                created_at TEXT,
                updated_at TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS role_permissions (
                role_id INTEGER,
                permission_id INTEGER,
                UNIQUE(role_id, permission_id)
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS user_roles (
                user_id INTEGER,
                role_id INTEGER,
                UNIQUE(user_id, role_id)
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS import_errors (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                batch_id INTEGER,
                row_no INTEGER,
                field TEXT,
                error_message TEXT,
                created_at TEXT
            )
        """)
        existing = {r["name"] for r in conn.execute("PRAGMA table_info(import_batches)").fetchall()}
        for name, col_type in {
            "target_table": "TEXT",
            "mapping_json": "TEXT",
            "preview_json": "TEXT",
            "source_columns_json": "TEXT",
            "saved_rows": "INTEGER DEFAULT 0",
        }.items():
            if name not in existing:
                conn.execute(f"ALTER TABLE import_batches ADD COLUMN {name} {col_type}")

        audit_cols = {r["name"] for r in conn.execute("PRAGMA table_info(audit_log)").fetchall()}
        for name, col_type in {
            "username": "TEXT", "table_name": "TEXT", "record_id": "INTEGER", "batch_id": "INTEGER",
            "created_at": "TEXT",
        }.items():
            if name not in audit_cols:
                conn.execute(f"ALTER TABLE audit_log ADD COLUMN {name} {col_type}")

        ts = now()
        for role, perms in DEFAULT_ROLES.items():
            conn.execute("INSERT OR IGNORE INTO roles (name, description, created_at, updated_at) VALUES (?, ?, ?, ?)", (role, f"Default {role} role", ts, ts))
            role_id = conn.execute("SELECT id FROM roles WHERE name=?", (role,)).fetchone()["id"]
            for perm in perms:
                conn.execute("INSERT OR IGNORE INTO permissions (name, description, created_at, updated_at) VALUES (?, ?, ?, ?)", (perm, perm.replace("_", " "), ts, ts))
                perm_id = conn.execute("SELECT id FROM permissions WHERE name=?", (perm,)).fetchone()["id"]
                conn.execute("INSERT OR IGNORE INTO role_permissions (role_id, permission_id) VALUES (?, ?)", (role_id, perm_id))
        conn.commit()


def permissions_for_role(role: str) -> set[str]:
    ensure_admin_foundation()
    if role == "admin":
        return set().union(*DEFAULT_ROLES.values())
    with db() as conn:
        rows = conn.execute(
            """
            SELECT p.name FROM permissions p
            JOIN role_permissions rp ON rp.permission_id=p.id
            JOIN roles r ON r.id=rp.role_id
            WHERE r.name=?
            """,
            (role,),
        ).fetchall()
    return {r["name"] for r in rows} or set(DEFAULT_ROLES.get(role, []))


def require_permission(request: Request, permission: str) -> None:
    role = role_from_request(request)
    if role != "admin" and permission not in permissions_for_role(role):
        raise HTTPException(status_code=403, detail=f"Permission required: {permission}")


def can_view_prices(request: Request) -> bool:
    return role_from_request(request) == "admin" or "view_prices" in permissions_for_role(role_from_request(request))


def redact_prices(rows: list[dict[str, Any]], request: Request) -> list[dict[str, Any]]:
    if can_view_prices(request):
        return rows
    redacted = []
    for row in rows:
        redacted.append({key: ("REDACTED" if key.lower() in PRICE_FIELDS else value) for key, value in row.items()})
    return redacted


def normalize(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def table_columns(conn: sqlite3.Connection, table: str) -> set[str]:
    return {row["name"] for row in conn.execute(f"PRAGMA table_info({quote_ident(table)})").fetchall()}


def quote_ident(name: str) -> str:
    if not re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", name):
        raise HTTPException(status_code=400, detail=f"Unsafe identifier: {name}")
    return f'"{name}"'


def read_upload(filename: str, content: bytes) -> pd.DataFrame:
    suffix = Path(filename).suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return pd.read_excel(io.BytesIO(content)).fillna("")
    if suffix == ".csv":
        return pd.read_csv(io.BytesIO(content)).fillna("")
    raise HTTPException(status_code=400, detail="Upload must be CSV or Excel")


def guess_mapping(columns: list[str], fields: list[str]) -> dict[str, str]:
    normalized_cols = {normalize(c).replace(" ", "_"): c for c in columns}
    aliases = {
        "name": ["client", "client_name", "hospital", "customer", "name"],
        "client_name": ["client", "client_name", "hospital", "customer"],
        "department_name": ["department", "department_name", "dept"],
        "pn": ["pn", "p/n", "item_code", "part_number", "part_no", "ref"],
        "item_code": ["pn", "p/n", "item_code", "part_number", "part_no", "ref"],
        "serial_number": ["serial", "serial_number", "s/n", "sn"],
        "call_no": ["call_no", "case_no", "service_call", "source_reference"],
        "po_no": ["po_no", "purchase_order", "po"],
        "qty": ["qty", "quantity"],
        "quantity": ["qty", "quantity"],
    }
    mapping = {}
    for field in fields:
        candidates = [field, *aliases.get(field, [])]
        for candidate in candidates:
            key = normalize(candidate).replace(" ", "_")
            if key in normalized_cols:
                mapping[field] = normalized_cols[key]
                break
    return mapping


def mapped_rows(df: pd.DataFrame, mapping: dict[str, str]) -> list[dict[str, Any]]:
    rows = []
    for idx, raw in df.iterrows():
        mapped = {field: raw[column] for field, column in mapping.items() if column in df.columns}
        rows.append({"row_no": int(idx) + 2, "raw": raw.to_dict(), "mapped": clean_row(mapped)})
    return rows


def clean_row(row: dict[str, Any]) -> dict[str, Any]:
    cleaned = {}
    for key, value in row.items():
        if pd.isna(value):
            value = ""
        if isinstance(value, str):
            value = value.strip()
        cleaned[key] = value
    return cleaned


def resolve_client(conn: sqlite3.Connection, client_name: str) -> int | None:
    wanted = normalize(client_name)
    if not wanted:
        return None
    for row in conn.execute("SELECT id, name FROM clients").fetchall():
        if normalize(row["name"]) == wanted:
            return int(row["id"])
    return None


def resolve_department(conn: sqlite3.Connection, client_id: int | None, department_name: str) -> int | None:
    if not client_id or not normalize(department_name):
        return None
    row = conn.execute(
        "SELECT id FROM departments WHERE client_id=? AND lower(department_name)=lower(?)",
        (client_id, department_name),
    ).fetchone()
    return int(row["id"]) if row else None


def validate_import_rows(target: str, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    config = IMPORT_TARGETS[target]
    with db() as conn:
        validated = []
        for row in rows:
            mapped = dict(row["mapped"])
            errors = []
            for field in config["required"]:
                if not str(mapped.get(field, "")).strip():
                    errors.append(f"{field} is required")
            existing_id = find_existing_id(conn, target, mapped)
            action = "skip_duplicate" if existing_id else "insert"
            if target in {"departments", "contacts", "equipment", "service_calls"}:
                client_id = resolve_client(conn, str(mapped.get("client_name", "")))
                if mapped.get("client_name") and not client_id:
                    errors.append("client_name was not found")
                mapped["client_id"] = client_id
                mapped["department_id"] = resolve_department(conn, client_id, str(mapped.get("department_name", "")))
            validated.append({
                **row,
                "mapped": mapped,
                "status": "valid" if not errors else "error",
                "errors": errors,
                "action": action,
                "existing_id": existing_id,
            })
        return validated


def find_existing_id(conn: sqlite3.Connection, target: str, mapped: dict[str, Any]) -> int | None:
    if target == "clients":
        wanted = normalize(mapped.get("name"))
        for row in conn.execute("SELECT id, name FROM clients").fetchall():
            if normalize(row["name"]) == wanted:
                return int(row["id"])
    if target == "departments":
        client_id = resolve_client(conn, str(mapped.get("client_name", "")))
        return resolve_department(conn, client_id, str(mapped.get("department_name", "")))
    if target == "contacts":
        email = normalize(mapped.get("email"))
        if email:
            row = conn.execute("SELECT id FROM contacts WHERE lower(email)=?", (email,)).fetchone()
            if row:
                return int(row["id"])
        client_id = resolve_client(conn, str(mapped.get("client_name", "")))
        row = conn.execute("SELECT id FROM contacts WHERE client_id=? AND lower(name)=lower(?)", (client_id, mapped.get("name"))).fetchone() if client_id else None
        return int(row["id"]) if row else None
    if target == "equipment" and mapped.get("serial_number"):
        row = conn.execute("SELECT id FROM equipment WHERE lower(serial_number)=lower(?)", (mapped.get("serial_number"),)).fetchone()
        return int(row["id"]) if row else None
    if target == "inventory_items":
        pn = mapped.get("pn") or mapped.get("item_code")
        row = conn.execute("SELECT id FROM inventory_items WHERE lower(pn)=lower(?)", (pn,)).fetchone() if pn else None
        return int(row["id"]) if row else None
    if target == "service_calls":
        ref = mapped.get("source_reference") or mapped.get("case_no") or mapped.get("call_no")
        row = conn.execute("SELECT id FROM service_calls WHERE lower(call_no)=lower(?) OR lower(parent_case_reference)=lower(?)", (ref, ref)).fetchone() if ref else None
        return int(row["id"]) if row else None
    return None


def insert_import_row(conn: sqlite3.Connection, target: str, row: dict[str, Any], batch_id: int, username: str) -> int | None:
    if row["status"] != "valid" or row["action"] == "skip_duplicate":
        return None
    table = IMPORT_TARGETS[target]["table"]
    allowed = table_columns(conn, table)
    data = dict(row["mapped"])
    if target == "inventory_items" and not data.get("pn"):
        data["pn"] = data.get("item_code")
    data.pop("client_name", None)
    data.pop("department_name", None)
    data.pop("source", None)
    data.pop("source_reference", None)
    data = {key: value for key, value in data.items() if key in allowed and value not in (None, "")}
    if "created_at" in allowed:
        data.setdefault("created_at", now())
    if "updated_at" in allowed:
        data.setdefault("updated_at", now())
    if not data:
        return None
    columns = list(data)
    placeholders = ", ".join("?" for _ in columns)
    sql = f"INSERT INTO {quote_ident(table)} ({', '.join(quote_ident(c) for c in columns)}) VALUES ({placeholders})"
    cur = conn.execute(sql, tuple(data[c] for c in columns))
    record_id = int(cur.lastrowid)
    audit(conn, username, "import_insert", table, record_id, batch_id, data)
    return record_id


def audit(conn: sqlite3.Connection, username: str, action: str, table: str, record_id: int | None, batch_id: int | None, data: Any) -> None:
    conn.execute(
        """
        INSERT INTO audit_log (username, action, table_name, record_id, batch_id, new_value, notes, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (username, action, table, record_id, batch_id, json.dumps(data, default=str), "admin foundation", now()),
    )


@router.get("/admin/database-map")
def database_map_page(request: Request):
    require_permission(request, "view_database_map")
    return FileResponse(STATIC_DIR / "database_map.html")


@router.get("/admin/imports")
def imports_page(request: Request):
    require_permission(request, "import_data")
    return FileResponse(STATIC_DIR / "imports.html")


@router.get("/administration/data-management")
@router.get("/admin/data-management")
def data_management_page(request: Request):
    require_permission(request, "data.import.preview")
    return FileResponse(STATIC_DIR / "data_management.html")


@router.get("/admin/query")
@router.get("/reports/query")
def query_page(request: Request):
    require_permission(request, "view_reports")
    if request.url.path == "/reports/query":
        return RedirectResponse("/admin/query", status_code=303)
    return FileResponse(STATIC_DIR / "query.html")


@router.get("/api/admin/security/me")
def security_me(request: Request):
    role = role_from_request(request)
    return {"username": username_from_request(request), "role": role, "permissions": sorted(permissions_for_role(role))}


@router.get("/api/admin/data-management/types")
def data_management_types(request: Request):
    require_permission(request, "data.import.preview")
    return [
        {
            "id": dataset.dataset_key,
            "label": dataset.display_name,
            "fields": dataset.field_names,
            "required": dataset.required_fields,
            "accepted_values": dataset.accepted_values,
        }
        for dataset in all_datasets()
    ]


@router.get("/api/admin/data-management/templates/{data_type}")
def download_template(data_type: str, request: Request):
    require_permission(request, "data.templates.download")
    try:
        dataset = get_dataset(data_type)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="Unknown data type")
    fields = dataset.field_names
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"{dataset.display_name} Import"[:31]
    sheet.append(fields)
    sheet.append(["required" if field in dataset.required_fields else "optional" for field in fields])

    instructions = workbook.create_sheet("Instructions")
    instructions.append(["Field", "Required", "Type", "Description", "Example", "Validation rule"])
    for field in dataset.fields:
        instructions.append([field.name, "yes" if field.required else "no", field.data_type, field.description, field.example, field.validation_rule])

    accepted = workbook.create_sheet("Accepted Values")
    accepted.append(["Field", "Accepted values"])
    for field in fields:
        accepted.append([field, ", ".join(dataset.accepted_values.get(field, []))])

    example = workbook.create_sheet("Example Data")
    example.append(fields)
    for row in dataset.example_rows:
        example.append([row.get(field, "") for field in fields])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"{data_type}-import-template.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/api/admin/data-management/validation-errors")
def data_management_validation_errors(request: Request):
    require_permission(request, "data.import.history")
    with db() as conn:
        rows = []
        if "data_validation_errors" in {r["name"] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}:
            rows.extend(
                dict(row)
                for row in conn.execute(
                    """
                    SELECT import_batch_id, import_row_id, field_name, raw_value, error_message, severity,
                           '' AS suggested_value, is_resolved, created_at
                    FROM data_validation_errors
                    ORDER BY id DESC
                    LIMIT 200
                    """
                ).fetchall()
            )
        rows.extend(
            {
                "import_batch_id": row["batch_id"],
                "import_row_id": None,
                "field_name": row["field"],
                "raw_value": "",
                "error_message": row["error_message"],
                "severity": "error",
                "suggested_value": "",
                "is_resolved": False,
                "created_at": row["created_at"],
                "row_no": row["row_no"],
            }
            for row in conn.execute("SELECT batch_id, row_no, field, error_message, created_at FROM import_errors ORDER BY id DESC LIMIT 200").fetchall()
        )
    return rows[:200]


@router.get("/api/admin/data-management/validation-errors/export")
def data_management_validation_errors_export(request: Request):
    rows = data_management_validation_errors(request)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Validation Errors"
    columns = ["import_batch_id", "row_no", "field_name", "raw_value", "error_message", "severity", "suggested_value", "is_resolved", "created_at"]
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="validation-errors.xlsx"'},
    )


@router.get("/api/admin/database-map")
def database_map(request: Request):
    require_permission(request, "view_database_map")
    info = masked_database_url()
    if is_postgres_url():
        engine = create_engine(DATABASE_URL, future=True, pool_pre_ping=True)
        inspector = inspect(engine)
        tables = []
        with engine.connect() as conn:
            for table in inspector.get_table_names():
                count = conn.execute(text(f'SELECT COUNT(*) FROM "{table}"')).scalar_one()
                tables.append({
                    "name": table,
                    "columns": inspector.get_columns(table),
                    "primary_keys": inspector.get_pk_constraint(table).get("constrained_columns", []),
                    "foreign_keys": inspector.get_foreign_keys(table),
                    "row_count": count,
                    "last_updated": None,
                    "explanation": TABLE_EXPLANATIONS.get(table, ""),
                })
        return {**info, "tables": tables}
    with db() as conn:
        rows = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name").fetchall()
        tables = []
        for row in rows:
            table = row["name"]
            cols = [dict(c) for c in conn.execute(f"PRAGMA table_info({quote_ident(table)})").fetchall()]
            pk = [c["name"] for c in cols if c.get("pk")]
            fks = [dict(fk) for fk in conn.execute(f"PRAGMA foreign_key_list({quote_ident(table)})").fetchall()]
            count = conn.execute(f"SELECT COUNT(*) AS c FROM {quote_ident(table)}").fetchone()["c"]
            last_updated = None
            col_names = {c["name"] for c in cols}
            if "updated_at" in col_names:
                last_updated = conn.execute(f"SELECT MAX(updated_at) AS v FROM {quote_ident(table)}").fetchone()["v"]
            elif "created_at" in col_names:
                last_updated = conn.execute(f"SELECT MAX(created_at) AS v FROM {quote_ident(table)}").fetchone()["v"]
            tables.append({
                "name": table, "columns": cols, "primary_keys": pk, "foreign_keys": fks,
                "row_count": count, "last_updated": last_updated, "explanation": TABLE_EXPLANATIONS.get(table, ""),
            })
        return {**info, "tables": tables}


@router.get("/api/admin/import-targets")
def import_targets(request: Request):
    require_permission(request, "import_data")
    return IMPORT_TARGETS


@router.post("/api/admin/imports/preview")
async def import_preview(
    request: Request,
    target: str = Form(...),
    file: UploadFile = File(...),
    mapping_json: str = Form(""),
):
    require_permission(request, "import_data")
    if target not in IMPORT_TARGETS:
        raise HTTPException(status_code=400, detail="Unsupported import target")
    content = await file.read()
    df = read_upload(file.filename or "upload", content)
    df = df.head(5000)
    columns = [str(c) for c in df.columns]
    mapping = json.loads(mapping_json) if mapping_json else guess_mapping(columns, IMPORT_TARGETS[target]["fields"])
    rows = validate_import_rows(target, mapped_rows(df, mapping))
    total = len(rows)
    valid = len([r for r in rows if r["status"] == "valid"])
    errors = total - valid
    with db() as conn:
        cur = conn.execute(
            """
            INSERT INTO import_batches
            (import_type, target_table, filename, status, total_rows, valid_rows, error_rows, created_by,
             created_at, mapping_json, preview_json, source_columns_json)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("admin_bulk_import", target, file.filename, "preview", total, valid, errors, username_from_request(request), now(), json.dumps(mapping), json.dumps(rows, default=str), json.dumps(columns)),
        )
        batch_id = int(cur.lastrowid)
        for r in rows:
            for error in r["errors"]:
                conn.execute("INSERT INTO import_errors (batch_id, row_no, error_message, created_at) VALUES (?, ?, ?, ?)", (batch_id, r["row_no"], error, now()))
        audit(conn, username_from_request(request), "import_preview", target, batch_id, batch_id, {"filename": file.filename, "rows": total})
        conn.commit()
    return {"batch_id": batch_id, "target": target, "columns": columns, "mapping": mapping, "summary": {"total_rows": total, "valid_rows": valid, "error_rows": errors}, "rows": rows[:100]}


@router.post("/api/admin/imports/{batch_id}/confirm")
def import_confirm(batch_id: int, request: Request):
    require_permission(request, "import_data")
    username = username_from_request(request)
    with db() as conn:
        batch = conn.execute("SELECT * FROM import_batches WHERE id=?", (batch_id,)).fetchone()
        if not batch:
            raise HTTPException(status_code=404, detail="Import batch not found")
        if batch["status"] not in {"preview", "validated"}:
            raise HTTPException(status_code=400, detail="Only preview batches can be confirmed")
        target = batch["target_table"]
        rows = json.loads(batch["preview_json"] or "[]")
        saved = 0
        skipped = 0
        for row in rows:
            record_id = insert_import_row(conn, target, row, batch_id, username)
            if record_id:
                saved += 1
            else:
                skipped += 1
        conn.execute(
            "UPDATE import_batches SET status=?, saved_rows=?, committed_at=? WHERE id=?",
            ("committed", saved, now(), batch_id),
        )
        audit(conn, username, "import_confirm", target, batch_id, batch_id, {"saved": saved, "skipped": skipped})
        conn.commit()
    return {"batch_id": batch_id, "saved_rows": saved, "skipped_rows": skipped, "status": "committed"}


@router.get("/api/admin/imports")
def list_admin_imports(request: Request):
    require_permission(request, "import_data")
    with db() as conn:
        rows = conn.execute("SELECT id, import_type, target_table, filename, status, total_rows, valid_rows, error_rows, saved_rows, created_by, created_at, committed_at FROM import_batches ORDER BY id DESC LIMIT 100").fetchall()
    return [dict(r) for r in rows]


@router.post("/admin/backups/create")
def create_backup_endpoint(request: Request):
    require_permission(request, "create_backup")
    return create_backup(username_from_request(request))


def create_backup(username: str = "system") -> dict[str, Any]:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    if is_postgres_url():
        filename = f"inventory_{ts}.dump"
        path = BACKUP_DIR / filename
        result = subprocess.run(["pg_dump", DATABASE_URL, "-Fc", "-f", str(path)], capture_output=True, text=True)
        if result.returncode != 0:
            raise HTTPException(status_code=500, detail=result.stderr or "pg_dump failed")
        zip_path = zip_backup(path)
    else:
        src = sqlite_path()
        if not src.exists():
            raise HTTPException(status_code=404, detail=f"SQLite database not found: {src}")
        path = BACKUP_DIR / f"inventory_{ts}.db"
        shutil.copy2(src, path)
        zip_path = zip_backup(path)
    prune_backups()
    with db() as conn:
        audit(conn, username, "backup_create", "database", None, None, {"backup": path.name, "zip": zip_path.name})
        conn.commit()
    return {"filename": path.name, "zip_filename": zip_path.name, "created_at": now(), "path": str(path), "zip_path": str(zip_path)}


def zip_backup(path: Path) -> Path:
    zip_path = path.with_suffix(path.suffix + ".zip")
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.write(path, arcname=path.name)
    return zip_path


def prune_backups(keep: int = 30) -> None:
    backups = sorted(BACKUP_DIR.glob("inventory_*"), key=lambda p: p.stat().st_mtime, reverse=True)
    stems_seen = set()
    grouped = []
    for path in backups:
        stem = path.name.replace(".db.zip", ".db").replace(".dump.zip", ".dump")
        if stem not in stems_seen:
            stems_seen.add(stem)
            grouped.append(stem)
    for stale in grouped[keep:]:
        for path in BACKUP_DIR.glob(stale + "*"):
            path.unlink(missing_ok=True)


@router.get("/admin/backups")
def backups(request: Request):
    require_permission(request, "create_backup")
    return RedirectResponse("/admin/database-map", status_code=303)


@router.get("/admin/backups/{filename}/download")
def backup_download(filename: str, request: Request):
    require_permission(request, "create_backup")
    if "/" in filename or "\\" in filename or not filename.startswith("inventory_"):
        raise HTTPException(status_code=400, detail="Invalid backup filename")
    path = BACKUP_DIR / filename
    if not path.exists():
        raise HTTPException(status_code=404, detail="Backup not found")
    return FileResponse(path, filename=filename)


def run_report_sql(sql: str, request: Request) -> list[dict[str, Any]]:
    with db() as conn:
        rows = [dict(r) for r in conn.execute(sql).fetchall()]
    return redact_prices(rows, request)


@router.get("/api/admin/reports")
def report_list(request: Request):
    require_permission(request, "view_reports")
    return [{"id": key, "label": key.replace("_", " ").title()} for key in REPORTS]


@router.get("/api/admin/reports/{report_id}")
def report_run(report_id: str, request: Request):
    require_permission(request, "view_reports")
    if report_id not in REPORTS:
        raise HTTPException(status_code=404, detail="Report not found")
    return {"report_id": report_id, "rows": run_report_sql(REPORTS[report_id], request)}


@router.get("/api/admin/reports/{report_id}/export")
def report_export(report_id: str, request: Request):
    require_permission(request, "view_reports")
    if report_id not in REPORTS:
        raise HTTPException(status_code=404, detail="Report not found")
    rows = run_report_sql(REPORTS[report_id], request)
    return excel_response(rows, f"{report_id}.xlsx")


@router.post("/api/admin/query")
async def run_query(request: Request):
    require_permission(request, "run_select_queries")
    payload = await request.json()
    sql = str(payload.get("sql") or "").strip()
    if not re.match(r"(?is)^select\b", sql) or re.search(r"(?is)\b(insert|update|delete|drop|alter|create|replace|pragma|attach|detach|vacuum)\b", sql):
        raise HTTPException(status_code=400, detail="Only read-only SELECT queries are allowed")
    if ";" in sql.rstrip(";"):
        raise HTTPException(status_code=400, detail="Only one SELECT statement is allowed")
    with db() as conn:
        rows = [dict(r) for r in conn.execute(sql).fetchmany(500)]
    return {"rows": redact_prices(rows, request), "limit": 500}


@router.post("/api/admin/query/export")
async def query_export(request: Request):
    require_permission(request, "run_select_queries")
    payload = await request.json()
    sql = str(payload.get("sql") or "").strip()
    if not re.match(r"(?is)^select\b", sql) or re.search(r"(?is)\b(insert|update|delete|drop|alter|create|replace|pragma|attach|detach|vacuum)\b", sql):
        raise HTTPException(status_code=400, detail="Only read-only SELECT queries are allowed")
    with db() as conn:
        rows = [dict(r) for r in conn.execute(sql).fetchmany(5000)]
    return excel_response(redact_prices(rows, request), "query_results.xlsx")


def excel_response(rows: list[dict[str, Any]], filename: str) -> Response:
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    columns = list(rows[0].keys()) if rows else ["No rows"]
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col) for col in columns])
    output = io.BytesIO()
    wb.save(output)
    return Response(
        output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
