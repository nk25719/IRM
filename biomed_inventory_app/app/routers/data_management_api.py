from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.config.database import DATA_ROOT
from app.database import SessionLocal, get_db
from app.data_management.template_registry import DatasetDefinition, all_datasets, get_dataset
from app.models.foundation import AuditEvent, DataValidationError, ImportBatch, ImportRow

router = APIRouter(prefix="/api/data-management", tags=["data-management"])

MAX_UPLOAD_BYTES = 10 * 1024 * 1024
ALLOWED_EXTENSIONS = {".csv", ".xlsx"}
UPLOAD_ROOT = DATA_ROOT / "imports"
EXPORT_ROOT = DATA_ROOT / "exports"
TEMPLATE_ROOT = DATA_ROOT / "templates"
for path in [UPLOAD_ROOT / "original", UPLOAD_ROOT / "processed", UPLOAD_ROOT / "error-reports", EXPORT_ROOT, TEMPLATE_ROOT, DATA_ROOT / "logs"]:
    path.mkdir(parents=True, exist_ok=True)

EXPORT_COLUMN_MAP = {
    "clients": {"client_name": "name"},
    "departments": {"client_name": None, "department_code": None},
    "contacts": {"client_name": None, "department_name": None, "contact_name": "name"},
    "equipment": {"client_name": None, "department_name": None, "asset_number": "asset_tag", "equipment_category": None},
    "inventory_items": {"item_code": "pn", "status": "active"},
}


class ExportPreviewRequest(BaseModel):
    dataset_key: str
    columns: list[str] | None = None
    filters: dict[str, Any] = Field(default_factory=dict)
    include_deleted: bool = False
    limit: int = Field(20, ge=1, le=100)


class ExportDownloadRequest(ExportPreviewRequest):
    format: str = Field("xlsx", pattern="^(xlsx|csv)$")


def _db() -> Session:
    return SessionLocal()


def _safe_name(value: str) -> str:
    stem = Path(value or "upload").stem
    stem = re.sub(r"[^A-Za-z0-9_.-]+", "-", stem).strip(".-")[:80] or "upload"
    return stem


def _dataset_payload(dataset: DatasetDefinition) -> dict:
    return {
        "dataset_key": dataset.dataset_key,
        "display_name": dataset.display_name,
        "domain": dataset.domain,
        "description": dataset.description,
        "version": dataset.version,
        "updated_at": dataset.updated_at,
        "fields": [
            {
                "name": field.name,
                "label": field.label,
                "data_type": field.data_type,
                "required": field.required,
                "description": field.description,
                "example": field.example,
                "validation_rule": field.validation_rule,
                "export_default": field.export_default and not field.sensitive,
            }
            for field in dataset.fields
        ],
        "required_fields": dataset.required_fields,
        "optional_fields": dataset.optional_fields,
        "accepted_values": dataset.accepted_values,
        "permission": dataset.permission,
        "import_supported": dataset.import_supported,
        "export_supported": dataset.export_supported,
        "supported_formats": ["xlsx", "csv"],
    }


def _template_workbook(dataset: DatasetDefinition) -> Workbook:
    workbook = Workbook()
    data = workbook.active
    data.title = "Data"
    data.append(dataset.field_names)

    instructions = workbook.create_sheet("Instructions")
    instructions.append(["Field", "Required", "Type", "Description", "Example", "Validation rule"])
    for field in dataset.fields:
        instructions.append([field.name, "yes" if field.required else "no", field.data_type, field.description, field.example, field.validation_rule])

    accepted = workbook.create_sheet("Accepted Values")
    accepted.append(["Field", "Accepted values"])
    for field in dataset.fields:
        accepted.append([field.name, ", ".join(dataset.accepted_values.get(field.name, ()))])

    examples = workbook.create_sheet("Example Data")
    examples.append(dataset.field_names)
    for row in dataset.example_rows or ({"_example": "Example only"},):
        examples.append([row.get(field, "") for field in dataset.field_names])
    return workbook


def _read_upload(file: UploadFile, content: bytes) -> pd.DataFrame:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Upload must be .xlsx or .csv")
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="Upload exceeds the 10 MB limit")
    if not content:
        raise HTTPException(status_code=400, detail="Upload is empty")
    try:
        if suffix == ".csv":
            dataframe = pd.read_csv(io.BytesIO(content)).fillna("")
        else:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            if workbook.sheetnames and workbook[workbook.sheetnames[0]].max_row < 1:
                raise HTTPException(status_code=400, detail="Upload is missing headers")
            dataframe = pd.read_excel(io.BytesIO(content), sheet_name=0).fillna("")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Unable to read uploaded file") from exc
    columns = [str(column).strip() for column in dataframe.columns]
    if not columns or all(not column for column in columns):
        raise HTTPException(status_code=400, detail="Upload is missing headers")
    if len(columns) != len(set(column.lower() for column in columns)):
        raise HTTPException(status_code=400, detail="Upload contains duplicate headers")
    dataframe.columns = columns
    return dataframe


def _normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().casefold()).strip("_")


def _mapping_for(dataset: DatasetDefinition, columns: list[str], supplied: dict[str, str] | None = None) -> dict[str, str]:
    supplied = supplied or {}
    normalized_columns = {_normalize_header(column): column for column in columns}
    mapping = {}
    aliases = {
        "manufacturer": ["brand", "maker", "manufacturer_name"],
        "model": ["model_number", "model_name"],
        "serial_number": ["serial", "s_n", "sn"],
        "client_name": ["hospital", "customer", "client"],
        "department_name": ["department", "dept"],
        "contact_name": ["name", "contact"],
    }
    for field in dataset.field_names:
        if supplied.get(field) in columns:
            mapping[field] = supplied[field]
            continue
        candidates = [field, *aliases.get(field, [])]
        for candidate in candidates:
            if _normalize_header(candidate) in normalized_columns:
                mapping[field] = normalized_columns[_normalize_header(candidate)]
                break
    return mapping


def _validate_rows(dataset: DatasetDefinition, dataframe: pd.DataFrame, mapping: dict[str, str], limit: int = 50) -> tuple[list[dict], list[dict]]:
    rows = []
    errors = []
    for index, raw in dataframe.head(limit).iterrows():
        row_number = int(index) + 2
        normalized = {field: raw[column] for field, column in mapping.items() if column in dataframe.columns}
        row_errors = []
        for field in dataset.required_fields:
            if not str(normalized.get(field, "")).strip():
                row_errors.append(
                    {
                        "row_number": row_number,
                        "field_name": field,
                        "raw_value": normalized.get(field, ""),
                        "error_code": "required",
                        "error_message": f"{field} is required",
                        "severity": "error",
                    }
                )
        status = "Ready" if not row_errors else "Error"
        rows.append({"row_number": row_number, "raw_data": raw.to_dict(), "normalized_data": normalized, "status": status, "issues": row_errors})
        errors.extend(row_errors)
    return rows, errors


def _audit(db: Session, event_type: str, entity_type: str, entity_id: str | None, new_values: dict | None, request: Request | None = None) -> None:
    db.add(
        AuditEvent(
            event_type=event_type,
            entity_type=entity_type,
            entity_id=entity_id,
            new_values=new_values,
            ip_address=request.client.host if request and request.client else None,
            user_agent=request.headers.get("user-agent") if request else None,
        )
    )


def _allowed_export_columns(dataset: DatasetDefinition, requested: list[str] | None) -> list[str]:
    allowed = set(dataset.export_fields)
    columns = requested or dataset.default_export_fields or dataset.export_fields
    invalid = [column for column in columns if column not in allowed]
    if invalid:
        raise HTTPException(status_code=400, detail=f"Unsupported export columns: {', '.join(invalid)}")
    return columns


def _table_columns(db: Session, table: str) -> set[str]:
    rows = db.execute(text(f'PRAGMA table_info("{table}")')).mappings().all()
    return {row["name"] for row in rows}


def _quote_identifier(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'


def _export_rows(db: Session, dataset: DatasetDefinition, columns: list[str], filters: dict[str, Any], limit: int | None = None) -> list[dict]:
    if not dataset.export_supported or not dataset.export_table:
        raise HTTPException(status_code=400, detail="Export is not implemented for this dataset")
    existing = _table_columns(db, dataset.export_table)
    column_map = EXPORT_COLUMN_MAP.get(dataset.dataset_key, {})
    selected = []
    for field in columns:
        table_column = column_map.get(field, field)
        if table_column and table_column in existing:
            selected.append((field, table_column))
    if not selected:
        return []
    clauses = []
    params = {}
    if "status" in existing and filters.get("status"):
        clauses.append("LOWER(status)=LOWER(:status)")
        params["status"] = str(filters["status"])
    if not filters.get("include_deleted") and "is_deleted" in existing:
        clauses.append("COALESCE(is_deleted, 0)=0")
    select_columns = [
        f"{_quote_identifier(table_column)} AS {_quote_identifier(field)}"
        for field, table_column in selected
    ]
    select_list = ", ".join(select_columns)
    table_name = _quote_identifier(dataset.export_table)
    sql = f"SELECT {select_list} FROM {table_name}"
    if clauses:
        sql += " WHERE " + " AND ".join(clauses)
    if dataset.export_order_by and dataset.export_order_by in existing:
        order_column = _quote_identifier(dataset.export_order_by)
        sql += f" ORDER BY {order_column}"
    if limit:
        sql += " LIMIT :limit"
        params["limit"] = limit
    return [dict(row) for row in db.execute(text(sql), params).mappings().all()]


@router.get("/datasets")
def datasets():
    return [_dataset_payload(dataset) for dataset in all_datasets()]


@router.get("/templates")
def templates():
    return [_dataset_payload(dataset) for dataset in all_datasets()]


@router.get("/templates/{dataset_key}")
def template_detail(dataset_key: str):
    try:
        return _dataset_payload(get_dataset(dataset_key))
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="Unknown dataset") from exc


@router.get("/templates/{dataset_key}/download")
def template_download(dataset_key: str):
    try:
        dataset = get_dataset(dataset_key)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="Unknown dataset") from exc
    workbook = _template_workbook(dataset)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{dataset.dataset_key}-import-template.xlsx"'},
    )


@router.get("/summary")
def summary(db: Session = Depends(get_db)):
    recent = db.query(ImportBatch).order_by(ImportBatch.id.desc()).limit(10).all()
    unresolved = db.query(DataValidationError).filter(DataValidationError.is_resolved.is_(False)).count()
    failed_rows = db.query(ImportRow).filter(ImportRow.processing_status.in_(["error", "failed"])).count()
    successful_rows = db.query(ImportBatch).with_entities(ImportBatch.successful_rows).all()
    return {
        "cards": [
            {"metric_key": "recent_imports", "label": "Recent Imports", "count": len(recent), "severity": "info", "description": "Recent staged import batches.", "action_url": "/administration/data-management#history"},
            {"metric_key": "rows_imported", "label": "Rows Imported", "count": sum(row[0] or 0 for row in successful_rows), "severity": "success", "description": "Rows confirmed by import batches.", "action_url": "/administration/data-management#history"},
            {"metric_key": "rows_with_errors", "label": "Rows With Errors", "count": failed_rows, "severity": "warning", "description": "Import rows currently marked failed or error.", "action_url": "/administration/data-management#validation"},
            {"metric_key": "unresolved_validation", "label": "Unresolved Validation Issues", "count": unresolved, "severity": "warning", "description": "Open data_validation_errors records.", "action_url": "/administration/data-management#validation"},
            {"metric_key": "available_templates", "label": "Available Templates", "count": len(all_datasets()), "severity": "info", "description": "Registered template definitions.", "action_url": "/administration/data-management#templates"},
        ],
        "data_health": data_quality(db),
    }


def data_quality(db: Session) -> list[dict]:
    metrics = []
    table_names = {row[0] for row in db.execute(text("SELECT name FROM sqlite_master WHERE type='table'")).all()}
    if "equipment_models" in table_names:
        cols = _table_columns(db, "equipment_models")
        if "manufacturer_id" in cols:
            metrics.append({"metric_key": "models_missing_manufacturer", "label": "Equipment models without manufacturer", "count": db.execute(text("SELECT COUNT(*) FROM equipment_models WHERE manufacturer_id IS NULL")).scalar_one(), "severity": "warning", "description": "Models not linked to canonical manufacturers.", "action_url": "/administration/master-data"})
        if "equipment_category_id" in cols:
            metrics.append({"metric_key": "models_missing_category", "label": "Equipment models without category", "count": db.execute(text("SELECT COUNT(*) FROM equipment_models WHERE equipment_category_id IS NULL")).scalar_one(), "severity": "warning", "description": "Models not linked to equipment categories.", "action_url": "/administration/master-data"})
    if "equipment" in table_names and "serial_number" in _table_columns(db, "equipment"):
        metrics.append({"metric_key": "equipment_missing_serial", "label": "Equipment without serial number", "count": db.execute(text("SELECT COUNT(*) FROM equipment WHERE serial_number IS NULL OR TRIM(serial_number)=''")).scalar_one(), "severity": "error", "description": "Installed equipment records missing serial numbers.", "action_url": "/administration/data-management"})
    if "data_validation_errors" in table_names:
        metrics.append({"metric_key": "unresolved_validation_errors", "label": "Unresolved validation errors", "count": db.query(DataValidationError).filter(DataValidationError.is_resolved.is_(False)).count(), "severity": "warning", "description": "Open validation errors requiring review.", "action_url": "/administration/data-management#validation"})
    metrics.append({"metric_key": "duplicate_equipment_candidates", "label": "Duplicate equipment candidates", "count": None, "severity": "unavailable", "description": "Duplicate detection rules are not implemented yet.", "action_url": ""})
    return metrics


@router.post("/imports")
async def create_import(
    request: Request,
    dataset_key: str = Form(...),
    import_mode: str = Form("validate_only"),
    mapping_json: str | None = Form(None),
    notes: str | None = Form(None),
    file: UploadFile = File(...),
):
    try:
        dataset = get_dataset(dataset_key)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="Unknown dataset") from exc
    if not dataset.import_supported:
        raise HTTPException(status_code=400, detail="Import is not implemented for this dataset")
    content = await file.read()
    dataframe = _read_upload(file, content)
    mapping = _mapping_for(dataset, list(dataframe.columns), json.loads(mapping_json) if mapping_json else None)
    rows, errors = _validate_rows(dataset, dataframe, mapping)
    safe_filename = f"{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}-{uuid.uuid4().hex[:12]}-{_safe_name(file.filename or 'upload')}{Path(file.filename or '').suffix.lower()}"
    storage_path = UPLOAD_ROOT / "original" / safe_filename
    storage_path.write_bytes(content)
    checksum = hashlib.sha256(content).hexdigest()

    db = _db()
    try:
        batch = ImportBatch(
            source_type=dataset.dataset_key,
            source_filename=_safe_name(file.filename or "upload"),
            source_checksum=checksum,
            status="validated" if not errors else "validation_failed",
            total_rows=len(dataframe),
            processed_rows=len(rows),
            successful_rows=sum(1 for row in rows if row["status"] == "Ready"),
            failed_rows=sum(1 for row in rows if row["status"] == "Error"),
            notes=json.dumps({"import_mode": import_mode, "notes": notes or "", "stored_as": safe_filename, "mapping": mapping}),
        )
        db.add(batch)
        db.flush()
        import_rows = []
        for row in rows:
            import_row = ImportRow(
                import_batch_id=batch.id,
                row_number=row["row_number"],
                raw_data=row["raw_data"],
                normalized_data=row["normalized_data"],
                processing_status=row["status"].casefold(),
                error_message="; ".join(issue["error_message"] for issue in row["issues"]) or None,
            )
            db.add(import_row)
            db.flush()
            import_rows.append(import_row)
            for issue in row["issues"]:
                issue_values = dict(issue)
                issue_values.pop("row_number", None)
                db.add(DataValidationError(import_batch_id=batch.id, import_row_id=import_row.id, **issue_values))
        _audit(db, "import_staged", "import_batch", str(batch.id), {"dataset": dataset.dataset_key, "row_count": len(dataframe)}, request)
        db.commit()
        return {
            "batch_id": batch.id,
            "status": batch.status,
            "summary": {"total_rows": batch.total_rows, "ready_rows": batch.successful_rows, "error_rows": batch.failed_rows, "warning_rows": 0},
            "columns": list(dataframe.columns),
            "mapping": mapping,
            "rows": rows,
        }
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


@router.get("/imports")
def imports(limit: int = Query(25, ge=1, le=100), offset: int = Query(0, ge=0), status: str | None = None, dataset: str | None = None):
    db = _db()
    try:
        query = db.query(ImportBatch)
        if status:
            query = query.filter(ImportBatch.status == status)
        if dataset:
            query = query.filter(ImportBatch.source_type == dataset)
        total = query.count()
        rows = query.order_by(ImportBatch.id.desc()).offset(offset).limit(limit).all()
        return {
            "items": [
                {
                    "id": row.id,
                    "filename": row.source_filename,
                    "dataset": row.source_type,
                    "status": row.status,
                    "started_at": row.started_at,
                    "completed_at": row.completed_at,
                    "uploaded_by": row.imported_by_id,
                    "total_rows": row.total_rows,
                    "processed_rows": row.processed_rows,
                    "successful_rows": row.successful_rows,
                    "failed_rows": row.failed_rows,
                }
                for row in rows
            ],
            "total": total,
            "limit": limit,
            "offset": offset,
        }
    finally:
        db.close()


@router.get("/imports/{batch_id}")
def import_detail(batch_id: int):
    db = _db()
    try:
        batch = db.get(ImportBatch, batch_id)
        if not batch:
            raise HTTPException(status_code=404, detail="Import batch not found")
        error_count = db.query(DataValidationError).filter(DataValidationError.import_batch_id == batch_id).count()
        return {
            "batch": {
                "id": batch.id,
                "filename": batch.source_filename,
                "dataset": batch.source_type,
                "status": batch.status,
                "started_at": batch.started_at,
                "completed_at": batch.completed_at,
                "uploaded_by": batch.imported_by_id,
                "total_rows": batch.total_rows,
                "processed_rows": batch.processed_rows,
                "successful_rows": batch.successful_rows,
                "failed_rows": batch.failed_rows,
            },
            "error_count": error_count,
            "timeline": {"started_at": batch.started_at, "completed_at": batch.completed_at},
        }
    finally:
        db.close()


@router.get("/imports/{batch_id}/rows")
def import_rows(batch_id: int, limit: int = Query(50, ge=1, le=200), offset: int = Query(0, ge=0)):
    db = _db()
    try:
        query = db.query(ImportRow).filter(ImportRow.import_batch_id == batch_id)
        total = query.count()
        rows = query.order_by(ImportRow.row_number).offset(offset).limit(limit).all()
        return {"items": [{"id": row.id, "row_number": row.row_number, "status": row.processing_status, "raw_data": row.raw_data, "normalized_data": row.normalized_data, "error_message": row.error_message, "warning_message": row.warning_message} for row in rows], "total": total, "limit": limit, "offset": offset}
    finally:
        db.close()


@router.get("/imports/{batch_id}/errors")
def import_errors(batch_id: int):
    return [row for row in validation_errors(batch_id=batch_id)["items"]]


@router.get("/imports/{batch_id}/error-report")
def import_error_report(batch_id: int):
    rows = import_errors(batch_id)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=["batch", "row", "field", "raw_value", "error_code", "message", "severity", "resolved"])
    writer.writeheader()
    for row in rows:
        writer.writerow({"batch": row["import_batch_id"], "row": row["row_number"], "field": row["field_name"], "raw_value": row["raw_value"], "error_code": row["error_code"], "message": row["error_message"], "severity": row["severity"], "resolved": row["is_resolved"]})
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers={"Content-Disposition": f'attachment; filename="import-{batch_id}-errors.csv"'})


@router.post("/imports/{batch_id}/validate")
def validate_import(batch_id: int):
    return {"batch_id": batch_id, "status": "already_validated", "detail": "Rows are validated during staging in this milestone."}


@router.post("/imports/{batch_id}/confirm")
def confirm_import(batch_id: int):
    db = _db()
    try:
        batch = db.get(ImportBatch, batch_id)
        if not batch:
            raise HTTPException(status_code=404, detail="Import batch not found")
        if batch.failed_rows:
            raise HTTPException(status_code=400, detail="Blocking validation errors prevent confirmation")
        batch.status = "ready_for_execution"
        db.commit()
        return {"batch_id": batch.id, "status": batch.status, "detail": "Final production-table import execution is intentionally deferred."}
    finally:
        db.close()


@router.post("/exports/preview")
def export_preview(payload: ExportPreviewRequest):
    try:
        dataset = get_dataset(payload.dataset_key)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="Unknown dataset") from exc
    columns = _allowed_export_columns(dataset, payload.columns)
    db = _db()
    try:
        rows = _export_rows(db, dataset, columns, payload.filters | {"include_deleted": payload.include_deleted}, payload.limit)
        return {"dataset": dataset.dataset_key, "columns": columns, "rows": rows, "row_count": len(rows)}
    finally:
        db.close()


@router.post("/exports/download")
def export_download(payload: ExportDownloadRequest, request: Request):
    preview = export_preview(payload)
    filename = f"{payload.dataset_key}-export-{datetime.now(timezone.utc).strftime('%Y%m%d')}.{payload.format}"
    db = _db()
    try:
        _audit(db, "export_downloaded", "dataset", payload.dataset_key, {"columns": preview["columns"], "filters": payload.filters, "row_count": preview["row_count"]}, request)
        db.commit()
    finally:
        db.close()
    if payload.format == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=preview["columns"])
        writer.writeheader()
        writer.writerows(preview["rows"])
        return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers={"Content-Disposition": f'attachment; filename="{filename}"'})
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Export"
    sheet.append(preview["columns"])
    for row in preview["rows"]:
        sheet.append([row.get(column, "") for column in preview["columns"]])
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/validation-errors")
def validation_errors(batch_id: int | None = None, status: str | None = None, severity: str | None = None, limit: int = Query(100, ge=1, le=500), offset: int = Query(0, ge=0)):
    db = _db()
    try:
        query = db.query(DataValidationError).outerjoin(ImportRow, DataValidationError.import_row_id == ImportRow.id)
        if batch_id is not None:
            query = query.filter(DataValidationError.import_batch_id == batch_id)
        if severity:
            query = query.filter(DataValidationError.severity == severity)
        if status == "open":
            query = query.filter(DataValidationError.is_resolved.is_(False))
        elif status == "resolved":
            query = query.filter(DataValidationError.is_resolved.is_(True))
        total = query.count()
        rows = query.order_by(DataValidationError.id.desc()).offset(offset).limit(limit).all()
        row_numbers = {
            row.id: row.row_number
            for row in db.query(ImportRow).filter(ImportRow.id.in_([error.import_row_id for error in rows if error.import_row_id])).all()
        }
        return {
            "items": [
                {
                    "id": row.id,
                    "import_batch_id": row.import_batch_id,
                    "import_row_id": row.import_row_id,
                    "row_number": row_numbers.get(row.import_row_id),
                    "field_name": row.field_name,
                    "raw_value": row.raw_value,
                    "error_code": row.error_code,
                    "error_message": row.error_message,
                    "severity": row.severity,
                    "is_resolved": row.is_resolved,
                    "resolved_by_id": row.resolved_by_id,
                    "resolved_at": row.resolved_at,
                    "created_at": row.created_at,
                }
                for row in rows
            ],
            "total": total,
            "limit": limit,
            "offset": offset,
        }
    finally:
        db.close()
