from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


COMPANY_NAME = "IRM - International Repair & Maintenance"
COMPANY_SUBTITLE = "Biomedical Equipment, Spare Parts, and Service Solutions"


def money(value: Any) -> float:
    try:
        return round(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def calculate_item_total(item: dict[str, Any]) -> float:
    quantity = money(item.get("quantity"))
    unit_price = money(item.get("unit_price"))
    discount_percent = money(item.get("discount_percent"))
    return round(quantity * unit_price * (1 - discount_percent / 100), 2)


def calculate_totals(items: list[dict[str, Any]], discount_amount: Any = 0, vat_rate: Any = 0) -> dict[str, float]:
    subtotal = round(sum(calculate_item_total(item) for item in items), 2)
    discount = money(discount_amount)
    taxable = max(0, subtotal - discount)
    vat = round(taxable * (money(vat_rate) / 100), 2)
    total = round(taxable + vat, 2)
    return {"subtotal": subtotal, "discount_amount": discount, "vat_amount": vat, "total_amount": total}


def build_excel(quotation: dict[str, Any], items: list[dict[str, Any]], client: dict[str, Any] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    ws.sheet_view.showGridLines = False

    widths = [14, 22, 42, 10, 14, 12, 14, 18, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    dark = "1F4E78"
    soft = "EAF2F8"
    line = Side(style="thin", color="C9D6E2")
    border = Border(left=line, right=line, top=line, bottom=line)

    ws.merge_cells("A1:I1")
    ws["A1"] = COMPANY_NAME
    ws["A1"].font = Font(size=18, bold=True, color=dark)
    ws.merge_cells("A2:I2")
    ws["A2"] = COMPANY_SUBTITLE
    ws["A2"].font = Font(color="667085")
    ws.merge_cells("A4:I4")
    ws["A4"] = "QUOTATION"
    ws["A4"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A4"].fill = PatternFill("solid", fgColor=dark)
    ws["A4"].alignment = Alignment(horizontal="center")

    client_name = (client or {}).get("name") or quotation.get("client_name") or f"Client #{quotation.get('client_id') or ''}"
    details = [
        ("Client", client_name, "Quotation No.", quotation.get("quotation_number") or quotation.get("quotation_no")),
        ("Contact", quotation.get("contact_name") or quotation.get("contact_id") or "", "Date", quotation.get("quotation_date") or quotation.get("quote_date")),
        ("Department", quotation.get("department_name") or quotation.get("department_id") or "", "Valid Until", quotation.get("valid_until")),
        ("Currency", quotation.get("currency") or "USD", "Status", quotation.get("status") or "draft"),
    ]
    row = 6
    for left_label, left_value, right_label, right_value in details:
        ws[f"A{row}"] = left_label
        ws[f"B{row}"] = left_value
        ws[f"F{row}"] = right_label
        ws[f"G{row}"] = right_value
        for cell in (f"A{row}", f"F{row}"):
            ws[cell].font = Font(bold=True, color=dark)
        row += 1

    table_row = 12
    headers = ["Code", "MPN", "Description", "Qty", "Unit Price", "Disc %", "Line Total", "Warranty", "Delivery"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(table_row, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for offset, item in enumerate(items, start=1):
        current = table_row + offset
        values = [
            item.get("item_code") or item.get("ref") or "",
            item.get("manufacturer_part_number") or "",
            item.get("description") or "",
            money(item.get("quantity") if item.get("quantity") is not None else item.get("qty")),
            money(item.get("unit_price")),
            money(item.get("discount_percent")),
            money(item.get("line_total") if item.get("line_total") is not None else item.get("total_price")),
            item.get("warranty") or "",
            item.get("delivery_time") or "",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(current, col, value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in (4, 5, 6, 7):
            ws.cell(current, col).number_format = '#,##0.00'

    totals = calculate_totals(items, quotation.get("discount_amount"), quotation.get("vat_rate"))
    total_row = table_row + len(items) + 3
    total_lines = [
        ("Subtotal", totals["subtotal"]),
        ("Discount", totals["discount_amount"]),
        (f"VAT {money(quotation.get('vat_rate')):.2f}%", totals["vat_amount"]),
        ("Total", totals["total_amount"]),
    ]
    for label, value in total_lines:
        ws[f"F{total_row}"] = label
        ws[f"G{total_row}"] = value
        ws[f"F{total_row}"].font = Font(bold=True, color=dark)
        ws[f"G{total_row}"].font = Font(bold=label == "Total")
        ws[f"G{total_row}"].number_format = '#,##0.00'
        total_row += 1

    terms_row = total_row + 2
    ws[f"A{terms_row}"] = "Terms"
    ws[f"A{terms_row}"].font = Font(bold=True, color=dark)
    for label, value in [
        ("Payment", quotation.get("payment_terms")),
        ("Delivery", quotation.get("delivery_terms")),
        ("Warranty", quotation.get("warranty_terms")),
        ("Notes", quotation.get("notes")),
    ]:
        terms_row += 1
        ws[f"A{terms_row}"] = label
        ws[f"B{terms_row}"] = value or ""
        ws.merge_cells(start_row=terms_row, start_column=2, end_row=terms_row, end_column=7)
    terms_row += 3
    ws[f"A{terms_row}"] = "Prepared by"
    ws[f"F{terms_row}"] = "Client approval"
    ws[f"A{terms_row + 2}"] = "Signature: ____________________"
    ws[f"F{terms_row + 2}"] = "Signature: ____________________"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _minimal_pdf_bytes(lines: list[str]) -> bytes:
    escaped = [line.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)") for line in lines]
    stream_lines = ["BT", "/F1 10 Tf", "50 790 Td"]
    for idx, line in enumerate(escaped[:58]):
        if idx:
            stream_lines.append("0 -14 Td")
        stream_lines.append(f"({line}) Tj")
    stream_lines.append("ET")
    stream = "\n".join(stream_lines)
    objects = [
        "<< /Type /Catalog /Pages 2 0 R >>",
        "<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        "<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
        "<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        f"<< /Length {len(stream.encode('latin-1', 'ignore'))} >>\nstream\n{stream}\nendstream",
    ]
    content = "%PDF-1.4\n"
    offsets = []
    for idx, obj in enumerate(objects, start=1):
        offsets.append(len(content.encode("latin-1")))
        content += f"{idx} 0 obj\n{obj}\nendobj\n"
    xref = len(content.encode("latin-1"))
    content += f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n"
    for offset in offsets:
        content += f"{offset:010d} 00000 n \n"
    content += f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF"
    return content.encode("latin-1", "ignore")


def group_title(group: dict[str, Any]) -> str:
    left = group.get("equipment_name") or group.get("model") or "Equipment"
    model = group.get("model") or ""
    serial = group.get("serial_number") or ""
    sr = group.get("service_report_number") or ""
    title = left if model and model in left else " - ".join(part for part in [left, model] if part)
    if serial:
        title += f" - S.N.{serial}"
    if sr:
        title += f" SR#: {sr}"
    return title


def grouped_items(items: list[dict[str, Any]], groups: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    groups = groups or []
    result = []
    used_ids = set()
    for group in groups:
        group_items = group.get("items") or [item for item in items if item.get("equipment_group_id") == group.get("id")]
        result.append({"title": group_title(group), "items": group_items})
        used_ids.update(item.get("id") for item in group_items)
    flat = [item for item in items if item.get("id") not in used_ids and not item.get("equipment_group_id")]
    if flat:
        result.append({"title": "General Items", "items": flat})
    if not result:
        result.append({"title": "General Items", "items": []})
    return result


def build_pdf(quotation: dict[str, Any], items: list[dict[str, Any]], client: dict[str, Any] | None = None, groups: list[dict[str, Any]] | None = None) -> bytes:
    client_name = (client or {}).get("name") or quotation.get("client_name") or f"Client #{quotation.get('client_id') or ''}"
    totals = calculate_totals(items, quotation.get("discount_amount"), quotation.get("vat_rate"))
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=18 * mm, leftMargin=18 * mm, topMargin=16 * mm, bottomMargin=18 * mm)
        styles = getSampleStyleSheet()
        normal = ParagraphStyle("CMMNormal", parent=styles["Normal"], fontSize=9, leading=11)
        small = ParagraphStyle("CMMSmall", parent=normal, fontSize=8, leading=10)
        title = ParagraphStyle("CMMTitle", parent=styles["Heading1"], fontSize=14, leading=16, spaceAfter=6)
        pn_style = ParagraphStyle("PN", parent=normal, leading=12)
        currency = quotation.get("currency") or "USD"
        offer_ref = quotation.get("quotation_number") or quotation.get("quotation_no") or ""
        phone = quotation.get("phone_number") or quotation.get("phone") or (client or {}).get("phone") or ""
        email = quotation.get("email") or (client or {}).get("contact_email") or ""
        sales_person = quotation.get("sales_person") or quotation.get("prepared_by") or "Nagham Kheir"

        story = []
        header = Table([
            [Paragraph("<b>CMM</b><br/><font size='8'>Clinical Medical Maintenance</font>", styles["Title"]), Paragraph("<b>Financial Offer</b>", title)],
        ], colWidths=[110 * mm, 48 * mm])
        header.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (1, 0), (1, 0), "RIGHT"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        story.extend([header, Spacer(1, 6)])

        details = [
            ["Customer", client_name, "Date", quotation.get("quotation_date") or quotation.get("quote_date") or ""],
            ["Offer Ref.", offer_ref, "Sales Person", sales_person],
            ["Phone Number", phone, "Email", email],
        ]
        detail_table = Table(details, colWidths=[26 * mm, 64 * mm, 28 * mm, 40 * mm])
        detail_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.extend([detail_table, Spacer(1, 12), Paragraph("<b>Financial Offer:</b>", title)])

        data = [["Qty", "Description", f"{currency} Unit Price", f"{currency} Tot. Price"]]
        spans = []
        row_index = 1
        for group in grouped_items(items, groups):
            data.append([Paragraph(f"<b>{group['title']}</b>", normal), "", "", ""])
            spans.append(row_index)
            row_index += 1
            for item in group["items"]:
                code = item.get("item_code") or item.get("manufacturer_part_number") or item.get("ref") or ""
                desc = item.get("description") or ""
                description = Paragraph(f"<b>P/N: {code}</b><br/>Description: {desc}", pn_style)
                quantity = money(item.get("quantity") if item.get("quantity") is not None else item.get("qty"))
                line_total = calculate_item_total(item)
                data.append([f"{quantity:g}", description, f"{money(item.get('unit_price')):,.2f}", f"{line_total:,.2f}"])
                row_index += 1

        table = Table(data, repeatRows=1, colWidths=[18 * mm, 90 * mm, 28 * mm, 30 * mm])
        table_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9D9D9")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#808080")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]
        for row in spans:
            table_style.extend([
                ("SPAN", (0, row), (-1, row)),
                ("BACKGROUND", (0, row), (-1, row), colors.HexColor("#BFBFBF")),
                ("FONTNAME", (0, row), (-1, row), "Helvetica-Bold"),
                ("ALIGN", (0, row), (-1, row), "LEFT"),
            ])
        table.setStyle(TableStyle(table_style))
        story.extend([table, Spacer(1, 10)])

        total_data = [
            [f"TOTAL BEFORE VAT {currency}", f"{totals['subtotal']:,.2f}"],
            [f"{money(quotation.get('vat_rate')):g}% VAT {currency}", f"{totals['vat_amount']:,.2f}"],
            [f"TOTAL {currency}", f"{totals['total_amount']:,.2f}"],
        ]
        total_table = Table(total_data, colWidths=[54 * mm, 30 * mm], hAlign="RIGHT")
        total_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#808080")),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#EDEDED")),
        ]))
        story.extend([total_table, Spacer(1, 12)])
        story.extend([
            Paragraph("<b>Conditions:</b>", normal),
            Paragraph(f"Validity: {quotation.get('valid_until') or ''}", normal),
            Paragraph(f"Payment terms: {quotation.get('payment_terms') or ''}", normal),
        ])

        class NumberedCanvas:
            def __init__(self, *args, **kwargs):
                from reportlab.pdfgen.canvas import Canvas

                self._canvas = Canvas(*args, **kwargs)
                self._saved_page_states = []

            def __getattr__(self, name):
                return getattr(self._canvas, name)

            def showPage(self):
                self._saved_page_states.append(dict(self._canvas.__dict__))
                self._canvas._startPage()

            def save(self):
                page_count = len(self._saved_page_states)
                for state in self._saved_page_states:
                    self._canvas.__dict__.update(state)
                    self._canvas._page_count = page_count
                    footer(self._canvas, doc)
                    self._canvas.showPage()
                self._canvas.save()

        def footer(canvas, doc_obj):
            canvas.saveState()
            canvas.setFont("Helvetica", 8)
            canvas.drawString(18 * mm, 10 * mm, "Financial Offer")
            canvas.drawCentredString(A4[0] / 2, 10 * mm, "CMM-SA-F-04-03-Edition01")
            page_count = getattr(canvas, "_page_count", doc_obj.page)
            canvas.drawRightString(A4[0] - 18 * mm, 10 * mm, f"Page {doc_obj.page} of {page_count}")
            canvas.restoreState()

        doc.build(story, canvasmaker=NumberedCanvas)
        return output.getvalue()
    except Exception:
        lines = [
            COMPANY_NAME,
            COMPANY_SUBTITLE,
            "QUOTATION",
            f"Client: {client_name}",
            f"Quotation No.: {quotation.get('quotation_number') or quotation.get('quotation_no') or ''}",
            f"Date: {quotation.get('quotation_date') or quotation.get('quote_date') or ''} Valid: {quotation.get('valid_until') or ''}",
            "",
            "Items:",
        ]
        for item in items:
            lines.append(f"{item.get('item_code') or ''} {item.get('description') or ''} qty {item.get('quantity') or item.get('qty') or ''} total {item.get('line_total') or item.get('total_price') or ''}")
        lines.extend(["", f"Subtotal: {totals['subtotal']:,.2f}", f"VAT: {totals['vat_amount']:,.2f}", f"Total: {totals['total_amount']:,.2f}", "", "Signature: ____________________"])
        return _minimal_pdf_bytes(lines)
